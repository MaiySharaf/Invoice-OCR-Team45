# backend/ocr_engine.py
import json
import re
import pytesseract
from pathlib import Path
from PIL import Image, ImageFilter, ImageEnhance
from openpyxl import load_workbook

try:
    from pdf2image import convert_from_path
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("PDF support not available. Install pdf2image for PDF support.")

# ──────────────────────────────────────────────────────────────
# TEAMMATE'S PREPROCESSING (ADDED)
# ──────────────────────────────────────────────────────────────

def preprocess_image(img: Image.Image) -> Image.Image:
    """
    Enhance image for better OCR accuracy:
      - Convert to grayscale
      - Sharpen edges
      - Increase contrast
    (Merged from teammate's ocr.py)
    """
    img = img.convert("L")                          # grayscale
    img = img.filter(ImageFilter.SHARPEN)           # sharpen
    img = ImageEnhance.Contrast(img).enhance(2.0)   # boost contrast
    img = ImageEnhance.Sharpness(img).enhance(2.0)
    return img

# ──────────────────────────────────────────────────────────────
# OCR FUNCTIONS (ENHANCED)
# ──────────────────────────────────────────────────────────────

def run_tesseract(image_path):
    """Handle both images and PDFs with preprocessing"""
    path = str(image_path)
    
    # Check if it's a PDF
    if path.lower().endswith('.pdf') and PDF_SUPPORT:
        try:
            images = convert_from_path(path, dpi=300)
            all_text = []
            for image in images:
                processed = preprocess_image(image)
                text = pytesseract.image_to_string(processed)
                all_text.append(text)
            return "\n".join(all_text)
        except Exception as e:
            print(f"PDF processing error: {e}")
            return ""
    
    # Regular image processing with preprocessing
    img = Image.open(path)
    processed = preprocess_image(img)
    return pytesseract.image_to_string(processed)

# ──────────────────────────────────────────────────────────────
# PATTERNS (EXISTING)
# ──────────────────────────────────────────────────────────────

DATE_PATTERN = re.compile(r"\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b")
INVOICE_NO_PATTERN = re.compile(
    r"(?:invoice\s*(?:no|number|#)\s*[:\-]?\s*)([A-Z0-9\-]+)", re.IGNORECASE
)
TOTAL_PATTERN = re.compile(
    r"(?:total\s*[$:]?\s*)([\d\s]+[\.,]\d{2})", re.IGNORECASE
)
TAX_PATTERN = re.compile(
    r"(?:tax|vat)\s*[$:]?\s*([\d\s]+[\.,]\d{2})", re.IGNORECASE
)
DISCOUNT_PATTERN = re.compile(
    r"(?:discount)\s*[$:]?\s*([\d\s]+[\.,]\d{2})", re.IGNORECASE
)
IBAN_PATTERN = re.compile(r"\bIBAN\s*[:\-]?\s*([A-Z0-9]+)\b", re.IGNORECASE)
AMOUNT_PATTERN = re.compile(r"\d[\d\s]*[\.,]\d{2}")
DUE_DATE_PATTERN = re.compile(
    r"(?:due\s*date\s*[:\-]?\s*)(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", re.IGNORECASE
)
SELLER_CLIENT_PATTERN = re.compile(
    r"Seller\s*:\s*(.*?)\s+Client\s*:\s*(.*?)\s+(?:\d|Tax\s*Id|IBAN|ITEMS)",
    re.IGNORECASE | re.DOTALL,
)

def normalize_space(value):
    return " ".join((value or "").replace("\n", " ").split())

def normalize_value(value):
    value = normalize_space(value)
    value = value.replace("$", "")
    value = re.sub(r"\s+", " ", value)
    money_like = re.fullmatch(r"[\d\s.,]+", value.strip())
    if money_like and ("," in value or "." in value):
        compact = value.replace(" ", "")
        last_dot = compact.rfind(".")
        last_comma = compact.rfind(",")
        if last_dot > last_comma:
            decimal_sep = "."
        else:
            decimal_sep = ","
        int_part, frac_part = compact.rsplit(decimal_sep, 1)
        int_part = int_part.replace(",", "").replace(".", "")
        value = f"{int_part}.{frac_part}"
    return value.strip().lower()

def load_label_rows(labels_path, max_samples=None):
    labels_path = Path(labels_path)
    if not labels_path.exists():
        raise FileNotFoundError(f"Labels not found at {labels_path}")

    print(f"Loading labels from {labels_path}...")   
    wb = load_workbook(labels_path, data_only=True)
    ws = wb.active
    rows = []
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        file_name = row[0]
        json_data = row[1]
        ocred_text = row[2]
        if not file_name or not json_data:
            continue
        try:
            label_obj = json.loads(json_data)
        except json.JSONDecodeError:
            continue
        rows.append({
            "file_name": str(file_name),
            "label": label_obj,
            "ocred_text": str(ocred_text or ""),
        })
        count += 1
        if max_samples and count >= max_samples:
            break
    return rows

load_label_rows_fixed = load_label_rows

def _cache_key_for_image(image_path):
    return f"{image_path.stem}.json"

def get_ocr_text_cached(image_path, cache_dir, refresh_cache=False):
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / _cache_key_for_image(image_path)
    if cache_path.exists() and not refresh_cache:
        try:
            payload = json.loads(cache_path.read_text(encoding="utf-8"))
            cached_text = payload.get("ocr_text", "")
            if isinstance(cached_text, str):
                return cached_text
        except json.JSONDecodeError:
            pass
    ocr_text = run_tesseract(image_path)
    cache_path.write_text(json.dumps({"ocr_text": ocr_text}, ensure_ascii=False), encoding="utf-8")
    return ocr_text

def build_default_output():
    return {
        "invoice": {
            "client_name": "", "client_address": "", "seller_name": "",
            "seller_address": "", "invoice_number": "", "invoice_date": "",
            "due_date": "",
        },
        "items": [],
        "item_count": "",
        "subtotal": {"tax": "", "discount": "", "total": ""},
        "payment_instructions": {
            "due_date": "", "bank_name": "", "account_number": "", "payment_method": "",
        },
    }
